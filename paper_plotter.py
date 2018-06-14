import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
import operator as op
import pandas as pd
"""
glide_y = [0.663866753,0.011526094,0.403690037,0.816768484]
vrocs_y = [0.530792501,0.022974083,0.354614412,0.682611997]
canvas_y = [0.539186058,0.01220381,0.112410451,0.745700043]

c_y = [0.662548743,0.141181543,0.43283818,0.806518438]
c_Ln_y = [0.708611218,0.102605514,0.544037084,0.825212109]
w_c_y = [0.663683441,0.139839837,0.444268858,0.80131062]
w_c_Ln_y = [0.709940451,0.099387002,0.561525495,0.822323462]
"""
from openpyxl import load_workbook


prop_cycle = plt.rcParams['axes.prop_cycle']
colors = prop_cycle.by_key()['color']
print(colors)
colors[5:7] = colors[3:5]
orange_red = "#d66926"
magenta = "#bd67b7"
colors[3] = orange_red
colors[4] = magenta
print(colors)
sns.set_palette(colors)
#sns.palplot(sns.color_palette())

# Load in the workbook
wb = load_workbook('auc_ef_proc.xlsx')

# Get sheet names
s_names = wb.sheetnames
print(s_names)

rows_to = [64,16,96]

glide_y = []
vrocs_y = []
canvas_y = []

for m in range(3):
    sheet = wb[s_names[m]]

    for i in range(rows_to[m]):
        if m == 0:
            glide_y.append(sheet.cell(row=i + 2, column=1).value)
        if m == 1:
            vrocs_y.append(sheet.cell(row=i + 2, column=1).value)
        if m == 2:
            canvas_y.append(sheet.cell(row=i + 2, column=1).value)


#c_y = [0.669649446,0.806518438,0.74118891,0.43283818]
#c_Ln_y = [0.720756458,0.825212109,0.744439223,0.544037084]
#w_c_y = [0.766995614,0.80131062,0.642158672,0.444268858]
#w_c_Ln_y = [0.773809524,0.822323462,0.682103321,0.561525495]
c_y = [0.729584423,0.808403461,0.728899714,0.560706343]
c_Ln_y = [0.732660079,0.819829323,0.756009743,0.629037382]
w_c_y = [0.731037489,0.807070049,0.672349889,0.553187546]
w_c_Ln_y = [0.739562143,0.820919758,0.70216033,0.629619824]


#ax = sns.boxplot(x="method", y="auc", data=[glide_y,vrocs_y,canvas_y,c_y,c_Ln_y,w_c_y,w_c_Ln_y])

#x = sns.swarmplot(x="method", y="auc", data=[glide_y,vrocs_y,canvas_y,c_y,c_Ln_y,w_c_y,w_c_Ln_y], color=".25")

#plt.boxplot([glide_y,vrocs_y,canvas_y,c_y,c_Ln_y,w_c_y,w_c_Ln_y])
idx_list_start = [0 , 16 ,20]
idx_list_stop = [15 , 19, 43]

#x = sns.swarmplot(x="method", y="auc", data=[glide_y,vrocs_y,canvas_y,c_y,c_Ln_y,w_c_y,w_c_Ln_y], color=".25")


col_list = ['Glide', 'Vrocs' , 'Canvas', 'Cons.' ,'Log Cons.' ,'Wght Cons.', 'Wght Log Cons.']
#plot_df = pd.DataFrame(columns='Glide',data=glide_y)
targets = ['CK1','DYRK1a', 'CDK5', 'GSK3b']
for i , t in enumerate(targets):
    plot_data =  {
        'Glide' : glide_y[i*16:(i+1)*16],
        'Vrocs' : vrocs_y[i*4:(i+1)*4],
        'Canvas' : canvas_y[i*24:(i+1)*24],
        'Cons.' : list([c_y[i]]),
        'Log Cons.' : list([c_Ln_y[i]]),
        'Wght Cons.' : list([w_c_y[i]]),
        'Wght Log Cons.' : list([w_c_Ln_y[i]])
    }

    #users_sorted_average = pd.DataFrame.from_dict(plot_data)
    #print(users_sorted_average.head())

    # sort keys and values together
    sorted_keys, sorted_vals = zip(*sorted(plot_data.items(), key=op.itemgetter(1)))
    #sorted_keys, sorted_vals = zip(plot_data.items(), key=op.itemgetter(1))

    #print(sorted_keys)
    #print(sorted_vals)
    # almost verbatim from question

    plt.figure(i)
    print(np.asarray(plot_data.values()))
    sns.set(context='notebook', style='whitegrid')
    sns.utils.axlabel(xlabel="Methods", ylabel="AUC", fontsize=12)
    #sns.boxplot(data=list(plot_data.values()), width=.25, whis = 10.0 )#, names = col_list)

    allkeys = ['Glide','Vrocs','Canvas','Cons.','Log Cons.','Wght Cons.','Wght Log Cons.']

    list_of_values = [plot_data.get(k) for k in allkeys if k ]


    sns.boxplot(data=list_of_values, width=.25, whis=10.0, color=sns.set_palette(colors))


    #sns.swarmplot(data=list_of_values, size=4, edgecolor="black", linewidth=.9, color=sns.set_palette(colors))

    plt.ylim([-0.00, 1.005])
    plt.xticks(rotation=30)
    # category labels
    plt.title(t)
    plt.xticks(plt.xticks()[0], allkeys)
    plt.subplots_adjust(left=None, bottom=0.2, right=None, top=None,
                    wspace=None, hspace=None)

    plt.gcf().savefig('saved_figures/box_plots/box_%s_%s' % (t, "no_swarm"))

    del(plot_data)
plt.show()